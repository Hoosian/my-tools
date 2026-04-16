"""Excel 生成器模块，提供表头、内容和数据生成功能。"""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from .random_data import generate


def create_workbook():
    """创建新的工作簿。"""
    return Workbook()


def add_headers(file_path: str, headers: list[str]):
    """在文件第一行添加表头。"""
    path = Path(file_path)

    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
        # 如果已有数据，在最前面插入一行
        ws.insert_rows(1)
    else:
        wb = create_workbook()
        ws = wb.active

    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    wb.save(path)
    return True


def add_content(file_path: str, rows: int, schema: dict[str, str]):
    """
    添加数据内容

    schema: dict[str, str] - 列名到数据类型的映射
    例如: {"姓名": "name", "年龄": "age", "邮箱": "email"}
    """
    path = Path(file_path)

    if not path.exists():
        wb = create_workbook()
        ws = wb.active
        start_row = 1
    else:
        wb = load_workbook(path)
        ws = wb.active
        start_row = ws.max_row + 1

    for i in range(rows):
        row_num = start_row + i
        for col, (header, data_type) in enumerate(schema.items(), start=1):
            if data_type == '_fixed_':
                value = header
            elif data_type == '_list_':
                value = header
            else:
                value = generate(data_type)
            ws.cell(row=row_num, column=col, value=value)

    wb.save(path)
    return True


def generate_excel(file_path: str, count: int, headers: list[str], schema: dict[str, str]):
    """生成完整的 Excel 文件。"""
    path = Path(file_path)
    wb = create_workbook()
    ws = wb.active

    # 添加表头
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # 添加数据
    for i in range(count):
        row_num = i + 2
        for col, (header, data_type) in enumerate(schema.items(), start=1):
            value = generate(data_type)
            ws.cell(row=row_num, column=col, value=value)

    wb.save(path)
    return True
